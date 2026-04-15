"""Exportacion a Excel con formato profesional."""

from io import BytesIO
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from engine.models import (
    SkuLine, ImportRates, CustomsTaxResult, ImportExpenses,
    SkuCostAllocation, SkuSellingPrice, ExportCostResult,
)


# Styles
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="0A6A72")
SUBHEADER_FONT = Font(name="Arial", bold=True, color="2B1F15", size=10)
SUBHEADER_FILL = PatternFill("solid", fgColor="D9EFF1")
NORMAL_FONT = Font(name="Arial", size=10)
MONEY_FONT = Font(name="Arial", size=10, color="054F56")
LABEL_FONT = Font(name="Arial", bold=True, size=10, color="705843")
THIN_BORDER = Border(
    left=Side(style="thin", color="D5C2A3"),
    right=Side(style="thin", color="D5C2A3"),
    top=Side(style="thin", color="D5C2A3"),
    bottom=Side(style="thin", color="D5C2A3"),
)
USD_FORMAT = '#,##0.00'
USD4_FORMAT = '#,##0.0000'
PCT_FORMAT = '0.00%'


def _style_header_row(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _style_data_row(ws, row, cols, formats=None):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = NORMAL_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center")
        if formats and col in formats:
            cell.number_format = formats[col]
            cell.font = MONEY_FONT


def _write_label_value(ws, row, label, value, fmt=None):
    ws.cell(row=row, column=1, value=label).font = LABEL_FONT
    ws.cell(row=row, column=1).border = THIN_BORDER
    cell = ws.cell(row=row, column=2, value=float(value) if isinstance(value, Decimal) else value)
    cell.font = MONEY_FONT
    cell.border = THIN_BORDER
    if fmt:
        cell.number_format = fmt


def _auto_width(ws, min_width=12, max_width=40):
    for col_cells in ws.columns:
        max_len = min_width
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[col_letter].width = max_len


def generate_excel(
    skus: list[SkuLine],
    rates: ImportRates,
    customs: CustomsTaxResult,
    expenses: ImportExpenses,
    allocations: list[SkuCostAllocation],
    prices: list[SkuSellingPrice],
    export_result: ExportCostResult | None,
    exchange_rate: Decimal,
) -> bytes:
    """Genera un workbook Excel completo con todas las hojas."""

    wb = Workbook()

    # --- Sheet 1: Factura Comercial ---
    ws = wb.active
    ws.title = "Factura Comercial"
    headers = ["Producto", "Unidad", "Cantidad", "Precio FOB Unit. (USD)",
               "FOB Total (USD)", "Proporcion (%)"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, 1, len(headers))

    formats = {4: USD4_FORMAT, 5: USD_FORMAT, 6: PCT_FORMAT}
    for i, s in enumerate(skus, 2):
        ws.cell(row=i, column=1, value=s.name)
        ws.cell(row=i, column=2, value=s.unit)
        ws.cell(row=i, column=3, value=int(s.quantity))
        ws.cell(row=i, column=4, value=float(s.fob_unit_price))
        ws.cell(row=i, column=5, value=float(s.fob_total))
        ws.cell(row=i, column=6, value=float(s.proportion))
        _style_data_row(ws, i, len(headers), formats)

    total_row = len(skus) + 2
    ws.cell(row=total_row, column=1, value="TOTAL").font = SUBHEADER_FONT
    ws.cell(row=total_row, column=3, value=int(sum(s.quantity for s in skus)))
    ws.cell(row=total_row, column=5, value=float(sum(s.fob_total for s in skus)))
    ws.cell(row=total_row, column=5).number_format = USD_FORMAT
    ws.cell(row=total_row, column=5).font = Font(name="Arial", bold=True, size=10, color="054F56")
    _auto_width(ws)

    # --- Sheet 2: Tributos Aduaneros ---
    ws2 = wb.create_sheet("Tributos Aduaneros")
    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 20

    ws2.cell(row=1, column=1, value="TRIBUTOS ADUANEROS")
    ws2.cell(row=1, column=1).font = Font(name="Arial", bold=True, size=14, color="0A6A72")
    ws2.merge_cells("A1:B1")

    data_rows = [
        ("FOB", customs.fob, USD_FORMAT),
        ("Flete Internacional", customs.freight, USD_FORMAT),
        ("CFR (FOB + Flete)", customs.cfr, USD_FORMAT),
        ("Seguro Internacional", customs.insurance, USD_FORMAT),
        ("Prima del Seguro", customs.insurance_premium, USD_FORMAT),
        ("CIF (CFR + Seguro)", customs.cif, USD_FORMAT),
        ("", "", None),
        ("Ad-Valorem", customs.ad_valorem, USD_FORMAT),
        ("ISC", customs.isc, USD_FORMAT),
        ("IGV Aduanero (16%)", customs.igv, USD_FORMAT),
        ("IPM (2%)", customs.ipm, USD_FORMAT),
        ("Total Tributos", customs.total_taxes, USD_FORMAT),
        ("", "", None),
        ("Percepcion IGV", customs.igv_perception, USD_FORMAT),
        ("Comision Agente de Aduanas", customs.broker_fee, USD_FORMAT),
        ("Deuda Aduanera Total", customs.total_deuda_aduanera, USD_FORMAT),
    ]
    for i, (label, value, fmt) in enumerate(data_rows, 3):
        if label:
            _write_label_value(ws2, i, label, value, fmt)

    # --- Sheet 3: Gastos Importacion ---
    ws3 = wb.create_sheet("Gastos Importacion")
    ws3.column_dimensions["A"].width = 40
    ws3.column_dimensions["B"].width = 18

    ws3.cell(row=1, column=1, value="GASTOS DE IMPORTACION (USD)")
    ws3.cell(row=1, column=1).font = Font(name="Arial", bold=True, size=14, color="0A6A72")
    ws3.merge_cells("A1:B1")

    for i, line in enumerate(expenses.lines, 3):
        _write_label_value(ws3, i, line.label, line.amount_usd, USD_FORMAT)

    sep_row = len(expenses.lines) + 4
    _write_label_value(ws3, sep_row, "Subtotal Gastos", expenses.subtotal_expenses, USD_FORMAT)
    _write_label_value(ws3, sep_row + 1, "IGV (18%) sobre Gastos", expenses.igv_on_expenses, USD_FORMAT)
    _write_label_value(ws3, sep_row + 2, "TOTAL GASTOS (inc. IGV)", expenses.total_expenses, USD_FORMAT)
    ws3.cell(row=sep_row + 2, column=2).font = Font(name="Arial", bold=True, size=11, color="054F56")

    # --- Sheet 4: Costo por Producto ---
    ws4 = wb.create_sheet("Costo por Producto")
    headers4 = ["Producto", "Unidad", "Cantidad", "FOB Total",
                "Proporcion", "CIF Asignado", "Tributos", "Gastos",
                "Costo Total (USD)", "Costo Unit. (USD)", "Costo Unit. (PEN)"]
    for col, h in enumerate(headers4, 1):
        ws4.cell(row=1, column=col, value=h)
    _style_header_row(ws4, 1, len(headers4))

    formats4 = {4: USD_FORMAT, 5: PCT_FORMAT, 6: USD_FORMAT, 7: USD_FORMAT,
                8: USD_FORMAT, 9: USD_FORMAT, 10: USD4_FORMAT, 11: USD4_FORMAT}
    for i, a in enumerate(allocations, 2):
        ws4.cell(row=i, column=1, value=a.sku.name)
        ws4.cell(row=i, column=2, value=a.sku.unit)
        ws4.cell(row=i, column=3, value=int(a.sku.quantity))
        ws4.cell(row=i, column=4, value=float(a.sku.fob_total))
        ws4.cell(row=i, column=5, value=float(a.sku.proportion))
        ws4.cell(row=i, column=6, value=float(a.allocated_cif))
        ws4.cell(row=i, column=7, value=float(a.allocated_taxes))
        ws4.cell(row=i, column=8, value=float(a.allocated_expenses))
        ws4.cell(row=i, column=9, value=float(a.total_cost))
        ws4.cell(row=i, column=10, value=float(a.unit_cost))
        ws4.cell(row=i, column=11, value=float(a.unit_cost * exchange_rate))
        _style_data_row(ws4, i, len(headers4), formats4)
    _auto_width(ws4)

    # --- Sheet 5: Precio de Venta ---
    ws5 = wb.create_sheet("Precio de Venta")
    headers5 = ["Producto", "Costo Unit. (USD)", "Margen (%)",
                "Precio sin IGV (USD)", "IGV (18%)", "Precio Final (USD)",
                "Precio Final (PEN)", "Ganancia Unit. (USD)"]
    for col, h in enumerate(headers5, 1):
        ws5.cell(row=1, column=col, value=h)
    _style_header_row(ws5, 1, len(headers5))

    formats5 = {2: USD4_FORMAT, 3: PCT_FORMAT, 4: USD4_FORMAT,
                5: USD4_FORMAT, 6: USD4_FORMAT, 7: USD4_FORMAT, 8: USD4_FORMAT}
    for i, p in enumerate(prices, 2):
        ws5.cell(row=i, column=1, value=p.sku_name)
        ws5.cell(row=i, column=2, value=float(p.unit_cost))
        ws5.cell(row=i, column=3, value=float(p.margin_percent))
        ws5.cell(row=i, column=4, value=float(p.retail_price_ex_igv))
        ws5.cell(row=i, column=5, value=float(p.igv_on_price))
        ws5.cell(row=i, column=6, value=float(p.retail_price_inc_igv))
        ws5.cell(row=i, column=7, value=float(p.retail_price_inc_igv * exchange_rate))
        ws5.cell(row=i, column=8, value=float(p.profit_per_unit))
        _style_data_row(ws5, i, len(headers5), formats5)
    _auto_width(ws5)

    # --- Sheet 6: Costos Exportacion ---
    if export_result:
        ws6 = wb.create_sheet("Costos Exportacion")
        ws6.column_dimensions["A"].width = 40
        ws6.column_dimensions["B"].width = 18

        ws6.cell(row=1, column=1, value="COSTOS DE EXPORTACION (USD)")
        ws6.cell(row=1, column=1).font = Font(name="Arial", bold=True, size=14, color="0A6A72")
        ws6.merge_cells("A1:B1")

        for i, line in enumerate(export_result.lines, 3):
            _write_label_value(ws6, i, line.label, line.amount_usd, USD_FORMAT)

        sep = len(export_result.lines) + 4
        _write_label_value(ws6, sep, "FOB Bruto", export_result.fob_total, USD_FORMAT)
        _write_label_value(ws6, sep + 1, "Total Gastos Exportacion", export_result.total_export_cost, USD_FORMAT)
        _write_label_value(ws6, sep + 2, "FOB Neto", export_result.fob_net, USD_FORMAT)

    # --- Sheet 7: Configuracion ---
    ws7 = wb.create_sheet("Configuracion")
    ws7.column_dimensions["A"].width = 35
    ws7.column_dimensions["B"].width = 18

    ws7.cell(row=1, column=1, value="TASAS Y PARAMETROS")
    ws7.cell(row=1, column=1).font = Font(name="Arial", bold=True, size=14, color="0A6A72")

    config_rows = [
        ("Ad-Valorem (%)", float(rates.ad_valorem_rate * 100), "0.00"),
        ("ISC (%)", float(rates.isc_rate * 100), "0.00"),
        ("IGV Aduanero (%)", float(rates.igv_rate * 100), "0.00"),
        ("IPM (%)", float(rates.ipm_rate * 100), "0.00"),
        ("Tasa Seguro (%)", float(rates.insurance_rate * 100), "0.00"),
        ("Prima Seguro (%)", float(rates.insurance_fee_rate * 100), "0.00"),
        ("Comision Agente (%)", float(rates.broker_commission_rate * 100), "0.000"),
        ("Minimo Agente (USD)", float(rates.broker_min_usd), USD_FORMAT),
        ("Percepcion IGV (%)", float(rates.igv_perception_rate * 100), "0.00"),
        ("IGV Estandar (%)", float(rates.standard_igv * 100), "0.00"),
        ("Tipo de Cambio (USD/PEN)", float(rates.exchange_rate), "0.0000"),
    ]
    for i, (label, value, fmt) in enumerate(config_rows, 3):
        _write_label_value(ws7, i, label, value, fmt)

    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
